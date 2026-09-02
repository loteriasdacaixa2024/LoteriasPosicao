# -*- coding: utf-8 -*-
"""Gera Excel da auditoria de menus e análises (5152–5160 + Central 8083)."""
from __future__ import annotations

import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart.series import SeriesLabel
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.series import SeriesLabel as _SL  # noqa: F401
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties as ShapeProps
from openpyxl.drawing.fill import ColorChoice as CC
from openpyxl.chart.series import SeriesLabel as SL2  # noqa: F401

OUT = Path(__file__).parent / "AUDITORIA_MENUS_ANALISES_MODALIDADES.xlsx"

# ---------------------------------------------------------------------------
# Dados da auditoria (código + HTTP 31/08/2026)
# ---------------------------------------------------------------------------
MODS = [
    ("lotofacil", "Lotofácil", 5152, "AnalisePorPosicao-Lotofacil-Only", "15 de 25 (P1–P15)"),
    ("diadesorte", "Dia de Sorte", 5153, "AnalisePorPosicao--DiaDeSorte-Only", "7 de 31 + Mês da Sorte"),
    ("lotomania", "Lotomania", 5154, "AnalisePorPosicao-Lotomania-Only", "20 sorteadas / aposta 50"),
    ("quina", "Quina", 5155, "AnalisePorPosicao-Quina-Only", "5 de 80"),
    ("megasena", "Mega-Sena", 5156, "AnalisePorPosicao-MegaSena-Only", "6 de 60"),
    ("maismilionaria", "+Milionária", 5157, "AnalisePorPosicao-MaisMilionaria-Only", "6 de 50 + 2 trevos"),
    ("duplasena", "Dupla Sena", 5158, "AnalisePorPosicao-DuplaSena-Only", "2 sorteios · 6 de 50"),
    ("timemania", "Timemania", 5159, "AnalisePorPosicao-Timemania-Only", "7 de 80 + Time do Coração"),
    ("supersete", "Super Sete", 5160, "AnalisePorPosicao-SuperSete-Only", "7 colunas 0–9"),
]
KEYS = [m[0] for m in MODS]
NOMES = {m[0]: m[1] for m in MODS}
SHORT = ["LF", "DDS", "LM", "QN", "MS", "+M", "DS", "TM", "SS"]

COLORS = {
    "lotofacil": "E8D7FA",
    "diadesorte": "B8E0C8",
    "lotomania": "FFF3CD",
    "quina": "CCE5FF",
    "megasena": "A8E6CF",
    "maismilionaria": "FFE0B2",
    "duplasena": "F8D7DA",
    "timemania": "D1ECF1",
    "supersete": "D4EDDA",
}

# Matriz: OK / 404 / AUSENTE / ESP / PARCIAL / EQUIV
# Ordem KEYS
MATRIZ = [
    ("Sniper / Frequência e Atrasos", ["OK*", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK"],
     "Página /analise/ (LF: /analise/atrasos). Frequência + atraso do volante."),
    ("Comparar concursos", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK"],
     "/analise/comparar-concursos/"),
    ("Repetição entre concursos", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK"],
     "/analise/repeticao-concursos/"),
    ("Resumo Geral da Modalidade", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK"],
     "/analise/resumo-geral/"),
    ("Resultados & Padrões", ["OK", "OK", "OK", "OK", "OK", "404", "404", "OK", "OK"],
     "/analise/analises-inteligentes/"),
    ("Análises Gerais", ["OK", "OK", "OK", "OK", "OK", "PARCIAL", "PARCIAL", "OK", "OK"],
     "Abas: Classificação, Dígitos, Diferencial Cruzado, Soma dígitos. +M/DS: 0 abas."),
    ("Análise de Somas e Dígitos", ["OK", "OK", "OK", "OK", "OK", "404", "404", "OK", "OK"],
     "/analise/somas-digitos/"),
    ("Linhas & DD × DU", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK"],
     "/analise/linhas-dd-du/"),
    ("Análise por Posição", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK"],
     "/analise/por-posicao/"),
    ("Concentração de Acertos", ["OK", "OK", "OK", "OK", "OK", "AUSENTE", "AUSENTE", "OK", "OK"],
     "Fora de CONCENTRACAO_MODALITIES em +M e DS."),
    ("Análise de Ciclos das Dezenas", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "ESP"],
     "Super Sete: spec enabled=False (motor colunas)."),
    ("Análise por Gaps e Ciclo", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK"],
     "/analise/gaps-ciclo/"),
    ("Análise Comportamental", ["OK", "OK", "OK", "OK", "OK", "404", "404", "OK", "OK"],
     "/analise/comportamento/"),
    ("Escolha Visual", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "404"],
     "Super Sete: menu injeta, wire ausente. Gerador já é False para colunas."),
    ("Análise Tubular", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "404"],
     "Idem Escolha Visual."),
    ("Central de Modelos + Backtesting", ["OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK", "OK"],
     "/modelos/"),
]

# Inventário de menu (título, grupo, href) — gerado da auditoria live
MENUS = {
    "lotofacil": {
        "Dados": [
            ("Sincronizar Dados", "/"),
            ("Comparar concursos", "/analise/comparar-concursos/"),
            ("Repetição entre concursos", "/analise/repeticao-concursos/"),
            ("Atrasos Posicionais", "/analise/atrasos"),
        ],
        "Análise": [
            ("Sniper por Posição", "/analise/atrasos"),
            ("Análise por Posição", "/analise/por-posicao/"),
            ("Resumo Geral da Modalidade", "/analise/resumo-geral/"),
            ("Resultados & Padrões", "/analise/analises-inteligentes/"),
            ("Análises Gerais", "/analise/analises-gerais/"),
            ("Análise de Somas e Dígitos", "/analise/somas-digitos/"),
            ("Linhas & DD × DU", "/analise/linhas-dd-du/"),
            ("Concentração de Acertos", "/analise/concentracao-acertos/"),
            ("Análise de Ciclos das Dezenas", "/analise/ciclo-cobertura/"),
            ("Análise por Gaps e Ciclo", "/analise/gaps-ciclo/"),
            ("Análise Comportamental", "/analise/comportamento/"),
            ("Escolha Visual", "/analise/escolha-visual/"),
            ("Análise Tubular", "/analise/analise-tubular/"),
            ("Comportamento LF", "/geradores-elite/comportamento-apostas/"),
            ("Central de Modelos", "/modelos/"),
            ("Backtesting Histórico", "/modelos/#pane-bt"),
        ],
        "Desdobramentos": [
            ("Desdobramento Inteligente", "/desdobramento/"),
            ("Lotofácil da Independência", "/desdobramento-especial/"),
        ],
        "Geradores de Elite": [
            ("Engine Final", "/geradores-elite/engine-final/"),
            ("Gerador Pro / GC", "/geradores-elite/gerador-gc/"),
            ("Gerador Elite", "/geradores-elite/gerador-elite/"),
            ("Repetição → Apostas", "/geradores-elite/repeticao-apostas/"),
            ("Ciclo — Apostas", "/geradores-elite/ciclo-apostas/"),
            ("Gaps e Ciclo → Apostas", "/geradores-elite/gaps-ciclo-apostas/"),
            ("Análise por Posição → Apostas", "/geradores-elite/gerador-por-posicao/"),
            ("Gerador por Concentração", "/geradores-elite/gerador-concentracao/"),
            ("Sniper → Apostas", "/geradores-elite/apostas-inteligentes/"),
            ("Comportamento → Apostas", "/geradores-elite/comportamento-apostas/"),
            ("Escolha/Tubular → Apostas", "/geradores-elite/escolha-tubular-apostas/"),
            ("Construtor de Construções", "/geradores-elite/construtor-construcoes/"),
            ("Gerador Inteligente por Dígitos", "/geradores-elite/gerador-digitos-inteligente/"),
        ],
    },
    "diadesorte": {
        "Dados": [
            ("Sincronizar Dados", "/"),
            ("Comparar concursos", "/analise/comparar-concursos/"),
            ("Repetição entre concursos", "/analise/repeticao-concursos/"),
            ("Análise Estatística", "/analise/"),
            ("Mês da Sorte", "/analise/#mes"),
        ],
        "Análise": [
            ("Sniper por Dezenas + Mês da Sorte", "/analise/"),
            ("Resumo Geral da Modalidade", "/analise/resumo-geral/"),
            ("Resultados & Padrões", "/analise/analises-inteligentes/"),
            ("Análises Gerais", "/analise/analises-gerais/"),
            ("Análise de Somas e Dígitos", "/analise/somas-digitos/"),
            ("Linhas & DD × DU", "/analise/linhas-dd-du/"),
            ("Análise por Posição", "/analise/por-posicao/"),
            ("Concentração de Acertos", "/analise/concentracao-acertos/"),
            ("Análise de Ciclos das Dezenas", "/analise/ciclo-cobertura/"),
            ("Análise por Gaps e Ciclo", "/analise/gaps-ciclo/"),
            ("Análise Comportamental", "/analise/comportamento/"),
            ("Escolha Visual", "/analise/escolha-visual/"),
            ("Análise Tubular", "/analise/analise-tubular/"),
            ("Central de Modelos", "/modelos/"),
            ("Backtesting Histórico", "/modelos/#pane-bt"),
        ],
        "Desdobramentos": [
            ("Desdobramentos", "/desdobramento/"),
        ],
        "Geradores de Elite": [
            ("Engine Final", "/geradores-elite/engine-final/"),
            ("Gerador Pro / GC", "/geradores-elite/gerador-gc/"),
            ("Gerador Elite", "/geradores-elite/gerador-elite/"),
            ("Repetição → Apostas", "/geradores-elite/repeticao-apostas/"),
            ("Ciclo — Apostas", "/geradores-elite/ciclo-apostas/"),
            ("Gaps e Ciclo → Apostas", "/geradores-elite/gaps-ciclo-apostas/"),
            ("Análise por Posição → Apostas", "/geradores-elite/gerador-por-posicao/"),
            ("Gerador por Concentração", "/geradores-elite/gerador-concentracao/"),
            ("Sniper → Apostas", "/geradores-elite/apostas-inteligentes/"),
            ("Comportamento → Apostas", "/geradores-elite/comportamento-apostas/"),
            ("Escolha/Tubular → Apostas", "/geradores-elite/escolha-tubular-apostas/"),
            ("Construtor de Construções", "/geradores-elite/construtor-construcoes/"),
            ("Gerador Inteligente por Dígitos", "/geradores-elite/gerador-digitos-inteligente/"),
        ],
    },
}

# Completar menus das outras a partir do padrão volante (auditoria live)
_VOLANTE_ANALISE = [
    ("Análise por Posição", "/analise/por-posicao/"),
    ("Resumo Geral da Modalidade", "/analise/resumo-geral/"),
    ("Resultados & Padrões", "/analise/analises-inteligentes/"),
    ("Análises Gerais", "/analise/analises-gerais/"),
    ("Análise de Somas e Dígitos", "/analise/somas-digitos/"),
    ("Linhas & DD × DU", "/analise/linhas-dd-du/"),
    ("Concentração de Acertos", "/analise/concentracao-acertos/"),
    ("Análise de Ciclos das Dezenas", "/analise/ciclo-cobertura/"),
    ("Análise por Gaps e Ciclo", "/analise/gaps-ciclo/"),
    ("Análise Comportamental", "/analise/comportamento/"),
    ("Escolha Visual", "/analise/escolha-visual/"),
    ("Análise Tubular", "/analise/analise-tubular/"),
]
_GE_FULL = MENUS["diadesorte"]["Geradores de Elite"]
_GE_SEM_CONC = [x for x in _GE_FULL if x[0] != "Gerador por Concentração"]
_GE_SS = [
    ("Engine Final", "/geradores-elite/engine-final/"),
    ("Gerador Pro / GC", "/geradores-elite/gerador-gc/"),
    ("Gerador Elite", "/geradores-elite/gerador-elite/"),
    ("Repetição → Apostas", "/geradores-elite/repeticao-apostas/"),
    ("Análise por Posição → Apostas", "/geradores-elite/gerador-por-posicao/"),
    ("Gerador por Concentração", "/geradores-elite/gerador-concentracao/"),
    ("Sniper → Apostas", "/geradores-elite/apostas-inteligentes/"),
    ("Comportamento → Apostas", "/geradores-elite/comportamento-apostas/"),
    ("Construtor de Construções", "/geradores-elite/construtor-construcoes/"),
    ("Gerador Inteligente por Dígitos", "/geradores-elite/gerador-digitos-inteligente/"),
    ("Gaps e Ciclo → Apostas", "/geradores-elite/gaps-ciclo-apostas/"),
]

MENUS["lotomania"] = {
    "Dados": [
        ("Sincronizar Dados", "/"),
        ("Comparar concursos", "/analise/comparar-concursos/"),
        ("Repetição entre concursos", "/analise/repeticao-concursos/"),
        ("Análise Estatística", "/analise/"),
        ("Repetição Consecutiva", "/analise/#repconsec"),
    ],
    "Análise": [("Sniper por Dezenas", "/analise/")] + _VOLANTE_ANALISE + [
        ("Comportamento LM", "/geradores-elite/comportamento-apostas/"),
        ("Central de Modelos", "/modelos/"),
        ("Backtesting Histórico", "/modelos/#pane-bt"),
    ],
    "Desdobramentos": [("Desdobramentos", "/desdobramento/")],
    "Geradores de Elite": list(_GE_FULL),
}
MENUS["quina"] = {
    "Dados": [
        ("Sincronizar Dados", "/"),
        ("Comparar concursos", "/analise/comparar-concursos/"),
        ("Repetição entre concursos", "/analise/repeticao-concursos/"),
        ("Análise Estatística", "/analise/"),
        ("Repetição Consecutiva", "/analise/#repconsec"),
    ],
    "Análise": [("Sniper por Dezenas", "/analise/")] + _VOLANTE_ANALISE + [
        ("Comportamento QN", "/geradores-elite/comportamento-apostas/"),
        ("Central de Modelos", "/modelos/"),
        ("Backtesting Histórico", "/modelos/#pane-bt"),
    ],
    "Desdobramentos": [
        ("Desdobramento Inteligente", "/desdobramento/"),
        ("Quina de São João", "/desdobramento-especial/"),
    ],
    "Geradores de Elite": list(_GE_FULL),
}
MENUS["megasena"] = {
    "Dados": [
        ("Sincronizar Dados", "/"),
        ("Comparar concursos", "/analise/comparar-concursos/"),
        ("Repetição entre concursos", "/analise/repeticao-concursos/"),
        ("Análise Estatística", "/analise/"),
        ("Repetição Consecutiva", "/analise/#repconsec"),
    ],
    "Análise": [("Sniper por Dezenas", "/analise/")] + _VOLANTE_ANALISE + [
        ("Comportamento MS", "/geradores-elite/comportamento-apostas/"),
        ("Central de Modelos", "/modelos/"),
        ("Backtesting Histórico", "/modelos/#pane-bt"),
    ],
    "Desdobramentos": [
        ("Des1 — Desdobramento Inteligente", "/desdobramento/"),
        ("Des2 — Desdobramento Estrutural", "/des2/"),
        ("Mega da Virada", "/desdobramento-especial/"),
    ],
    "Geradores de Elite": list(_GE_FULL),
}
_VOLANTE_SEM_CONC = [x for x in _VOLANTE_ANALISE if x[0] != "Concentração de Acertos"]
MENUS["maismilionaria"] = {
    "Dados": [
        ("Sincronizar Dados", "/"),
        ("Comparar concursos", "/analise/comparar-concursos/"),
        ("Repetição entre concursos", "/analise/repeticao-concursos/"),
        ("Análise Estatística", "/analise/"),
        ("Análise de Trevos", "/analise/#trevo"),
    ],
    "Análise": [("Sniper por Dezenas + Trevo", "/analise/")] + _VOLANTE_SEM_CONC + [
        ("Comportamento +M", "/geradores-elite/comportamento-apostas/"),
        ("Central de Modelos", "/modelos/"),
        ("Backtesting Histórico", "/modelos/#pane-bt"),
    ],
    "Desdobramentos": [("Desdobramento Inteligente", "/desdobramento/")],
    "Geradores de Elite": list(_GE_SEM_CONC),
}
MENUS["duplasena"] = {
    "Dados": [
        ("Sincronizar Dados", "/"),
        ("Comparar concursos", "/analise/comparar-concursos/"),
        ("Repetição entre concursos", "/analise/repeticao-concursos/"),
        ("Análise Estatística", "/analise/"),
        ("2º Sorteio", "/analise/#sorteio2"),
    ],
    "Análise": [("Sniper por Dezenas", "/analise/")] + _VOLANTE_SEM_CONC + [
        ("Comportamento DS2", "/geradores-elite/comportamento-apostas/"),
        ("Central de Modelos", "/modelos/"),
        ("Backtesting · Prêmio Duplo", "/modelos/#pane-bt"),
    ],
    "Desdobramentos": [
        ("Desdobramento Inteligente", "/desdobramento/"),
        ("Dupla de Páscoa", "/desdobramento-especial/"),
    ],
    "Geradores de Elite": list(_GE_SEM_CONC),
}
MENUS["timemania"] = {
    "Dados": [
        ("Sincronizar Dados", "/"),
        ("Comparar concursos", "/analise/comparar-concursos/"),
        ("Repetição entre concursos", "/analise/repeticao-concursos/"),
        ("Análise Estatística", "/analise/"),
        ("Ranking de Times", "/analise/#times"),
    ],
    "Análise": [("Sniper por Dezenas + Timemania", "/analise/")] + _VOLANTE_ANALISE + [
        ("Comportamento TM", "/geradores-elite/comportamento-apostas/"),
        ("Central de Modelos", "/modelos/"),
        ("Backtesting · 8 prêmios", "/modelos/#pane-bt"),
    ],
    "Desdobramentos": [("Desdobramentos", "/desdobramento/")],
    "Geradores de Elite": list(_GE_FULL),
}
_SS_ANALISE = [
    ("Análise por Posição", "/analise/por-posicao/"),
    ("Resumo Geral da Modalidade", "/analise/resumo-geral/"),
    ("Resultados & Padrões", "/analise/analises-inteligentes/"),
    ("Análises Gerais", "/analise/analises-gerais/"),
    ("Análise de Somas e Dígitos", "/analise/somas-digitos/"),
    ("Linhas & DD × DU", "/analise/linhas-dd-du/"),
    ("Concentração de Acertos", "/analise/concentracao-acertos/"),
    ("Análise por Gaps e Ciclo", "/analise/gaps-ciclo/"),
    ("Análise Comportamental", "/analise/comportamento/"),
    ("Escolha Visual", "/analise/escolha-visual/"),
    ("Análise Tubular", "/analise/analise-tubular/"),
]
MENUS["supersete"] = {
    "Dados": [
        ("Sincronizar Dados", "/"),
        ("Comparar concursos", "/analise/comparar-concursos/"),
        ("Repetição entre concursos", "/analise/repeticao-concursos/"),
        ("Análise Estatística", "/analise/"),
        ("Repetição de Dígitos", "/analise/#repeticoes"),
    ],
    "Análise": [("Sniper por Colunas", "/analise/")] + _SS_ANALISE + [
        ("Comportamento SS", "/geradores-elite/comportamento-apostas/"),
        ("Central de Modelos", "/modelos/"),
        ("Backtesting Histórico", "/modelos/#pane-bt"),
    ],
    "Desdobramentos": [("Desdobramentos", "/desdobramento/")],
    "Geradores de Elite": list(_GE_SS),
}

HTTP_MENU = {
    # path (sem hash) -> status por key; default 200 se omitido
    "/geradores-elite/repeticao-apostas/": {"lotofacil": 500},
    "/analise/analises-inteligentes/": {"maismilionaria": 404, "duplasena": 404},
    "/analise/somas-digitos/": {"maismilionaria": 404, "duplasena": 404},
    "/analise/comportamento/": {"maismilionaria": 404, "duplasena": 404},
    "/analise/concentracao-acertos/": {"maismilionaria": 404, "duplasena": 404},
    "/analise/escolha-visual/": {"supersete": 404},
    "/analise/analise-tubular/": {"supersete": 404},
    "/analise/ciclo-cobertura/": {"supersete": 404},
    "/geradores-elite/escolha-tubular-apostas/": {
        "maismilionaria": 500, "duplasena": 500, "supersete": 404,
    },
    "/geradores-elite/gerador-concentracao/": {
        "maismilionaria": 404, "duplasena": 404,
    },
    "/geradores-elite/ciclo-apostas/": {"supersete": 404},
    "/des2/": {k: 404 for k in KEYS if k != "megasena"},
    "/analise/atrasos": {k: 404 for k in KEYS if k != "lotofacil"},
}

QUEBRADOS = [
    ("+Milionária", "Análise", "Resultados & Padrões", "NÃO", "NÃO (404)", "SIM (shared)",
     "No menu, sem wire_analise_inteligentes", "CRÍTICO"),
    ("+Milionária", "Análise", "Análise de Somas e Dígitos", "NÃO", "NÃO (404)", "SIM",
     "No menu, sem wire_analise_somas_digitos", "CRÍTICO"),
    ("+Milionária", "Análise", "Análise Comportamental", "NÃO", "NÃO (404)", "SIM",
     "No menu, sem wire_analise_comportamento", "CRÍTICO"),
    ("Dupla Sena", "Análise", "Resultados & Padrões", "NÃO", "NÃO (404)", "SIM",
     "No menu, sem wire_analise_inteligentes", "CRÍTICO"),
    ("Dupla Sena", "Análise", "Análise de Somas e Dígitos", "NÃO", "NÃO (404)", "SIM",
     "No menu, sem wire_analise_somas_digitos", "CRÍTICO"),
    ("Dupla Sena", "Análise", "Análise Comportamental", "NÃO", "NÃO (404)", "SIM",
     "No menu, sem wire_analise_comportamento", "CRÍTICO"),
    ("Super Sete", "Análise", "Escolha Visual", "NÃO", "NÃO (404)", "SIM",
     "Menu injeta; tem_gerador_escolha_tubular=False. Avaliar se é específica de colunas.", "CRÍTICO"),
    ("Super Sete", "Análise", "Análise Tubular", "NÃO", "NÃO (404)", "SIM",
     "Idem Escolha Visual", "CRÍTICO"),
    ("+Milionária", "Geradores", "Escolha/Tubular → Apostas", "PARCIAL", "SIM (500)", "SIM",
     "Item no menu com erro de servidor", "ALTO"),
    ("Dupla Sena", "Geradores", "Escolha/Tubular → Apostas", "PARCIAL", "SIM (500)", "SIM",
     "Item no menu com erro de servidor", "ALTO"),
    ("Lotofácil", "Geradores", "Repetição → Apostas", "PARCIAL", "SIM (500)", "SIM",
     "Item no menu com erro de servidor", "ALTO"),
    ("Super Sete", "Análise", "Ciclos das Dezenas", "spec enabled=False", "NÃO", "SIM",
     "Específica da modalidade (motor colunas) — não é gap", "—"),
    ("+Milionária", "Análise", "Concentração de Acertos", "NÃO (spec)", "NÃO", "SIM",
     "Ausente de ponta a ponta (não está no menu)", "ALTO"),
    ("Dupla Sena", "Análise", "Concentração de Acertos", "NÃO (spec)", "NÃO", "SIM",
     "Ausente de ponta a ponta (não está no menu)", "ALTO"),
]

INEXISTENTES = [
    ("Concentração de Acertos", "+Milionária, Dupla Sena",
     "Fora de CONCENTRACAO_MODALITIES", "Implementar spec ou documentar como não aplicável"),
    ("Ciclo de Cobertura", "Super Sete",
     "ciclo_cobertura.specs enabled=False, motor=colunas", "Não tratar como falta"),
    ("Abas de Análises Gerais", "+Milionária, Dupla Sena",
     "_MODS_ESTUDOS omite as duas; ESTUDOS_MODALITIES inclui", "Incluir no registry"),
    ("Repetição Consecutiva (módulo shared)",
     "Lotofácil, Dia de Sorte, +Milionária, Dupla Sena, Timemania, Super Sete",
     "Só Quina/Lotomania usam o shared; Mega-Sena tem cópia local", "Replicar ou unificar"),
]

NOMENCLATURA = [
    ("Sniper por Dezenas / por Posição / por Colunas / + Mês / + Trevo / + Timemania",
     "Sniper", "Sufixo só do extra da modalidade (mês, trevo, time, colunas)"),
    ("Análise Estatística / Atrasos Posicionais",
     "Frequência e Atrasos", "Lotofácil usa outro rótulo e /analise/atrasos"),
    ("Comportamento LF / LM / QN / MS / +M / DS2 / TM / SS",
     "(remover do menu Análise)", "É o gerador; DDS já não duplica"),
    ("Análise Comportamental", "Análise Comportamental", "Manter"),
    ("Comportamento → Apostas", "Comportamento → Apostas", "Somente em Geradores de Elite"),
    ("Backtesting Histórico / · Prêmio Duplo / · 8 prêmios",
     "Backtesting", "Sufixo só quando a regra muda (Dupla, Timemania)"),
    ("Desdobramentos / Des1 / Des2 / concurso especial",
     "Desdobramento Inteligente + extras nomeados", "Mega-Sena é o modelo de dropdown"),
    ("Central de Modelos 5 MODELOS / 6 MODELOS / 6",
     "Central de Modelos", "Unificar badge"),
]

ROTAS = [
    ("Menu → rota não registrada", "+Milionária e Dupla Sena (3 análises); Super Sete (2)", "404"),
    ("Gerador no menu com erro", "Lotofácil repeticao-apostas; +M/DS escolha-tubular-apostas", "500"),
    ("/analise/atrasos", "Só Lotofácil implementa", "404 nas outras (esperado se digitada)"),
    ("/des2/", "Só Mega-Sena", "404 nas outras (esperado)"),
    ("/desdobramento-especial/", "LF, QN, MS, DS", "404 nas demais (esperado)"),
    ("/geradores-elite/ciclo-apostas/ Super Sete", "Fora do menu; ciclo desligado", "404 esperado"),
    ("Central :8083/configuracoes/", "Timeout na sonda HTTP", "Investigar à parte"),
    ("verify_nav.py", "Espera 3–4 geradores; menu real tem 11–13", "Checker obsoleto"),
    ("Mega-Sena sem register_conferencia_extras", "As outras 8 chamam", "Menu igual; APIs extras podem faltar"),
    ("Análise Estatística e Sniper", "Mesma URL /analise/ (LF: ambos /analise/atrasos)", "Duplicação de destino"),
    ("Comportamento XX no Análise = Comportamento → Apostas", "Mesma URL do gerador", "Duplicação de destino"),
]

ESPECIFICAS = [
    ("Mês da Sorte", "Dia de Sorte", "Extra do sorteio"),
    ("Análise de Trevos", "+Milionária", "Extra do sorteio"),
    ("Ranking de Times", "Timemania", "Extra do sorteio"),
    ("2º Sorteio", "Dupla Sena", "Dois sorteios por concurso"),
    ("Atrasos posicionais P1–P15 / Sniper por Posição", "Lotofácil", "15 posições ordenadas"),
    ("Sniper por Colunas / Repetição de Dígitos", "Super Sete", "7 colunas 0–9"),
    ("Ciclo de Cobertura desligado", "Super Sete", "Spec: motor colunas"),
    ("Escolha Visual / Tubular (não implementar)", "Super Sete",
     "tem_gerador_escolha_tubular=False; o menu Análise ainda injeta — isso é bug"),
    ("Independência / São João / Virada / Páscoa", "LF, QN, MS, DS", "Concursos especiais"),
    ("Des2 estrutural", "Mega-Sena", "Módulo exclusivo"),
]

PRIORIDADE = [
    ("CRÍTICO", "+Milionária e Dupla Sena",
     "Registrar wire_analise_inteligentes, wire_analise_somas_digitos, wire_analise_comportamento — ou tirar do menu",
     "Preserva o que já funciona; só completa o que o menu promete"),
    ("CRÍTICO", "Super Sete",
     "Tirar Escolha Visual e Análise Tubular do menu Análise, ou implementar (há argumento para não implementar)",
     "Menu não deve apontar para 404"),
    ("ALTO", "+Milionária e Dupla Sena",
     "Incluir em _MODS_ESTUDOS (Análises Gerais abre sem abas)",
     "Rota 200 com conteúdo vazio"),
    ("ALTO", "Lotofácil / +M / DS",
     "Corrigir HTTP 500 dos geradores no menu",
     "Item visível quebrado"),
    ("ALTO", "+Milionária e Dupla Sena",
     "Decidir Concentração: implementar spec ou documentar como não aplicável",
     "Não está no menu hoje"),
    ("MÉDIO", "Todas menos DDS",
     "Remover Comportamento XX do menu Análise (ficar só Análise Comportamental + gerador em Elite)",
     "Padronizar com Dia de Sorte"),
    ("MÉDIO", "Todas",
     "Unificar rótulos de Sniper e Backtesting",
     "Só texto"),
    ("MÉDIO", "Lotofácil",
     "Se /analise/ for página distinta, expor Frequência e Atrasos além de Atrasos Posicionais",
     "Verificar se há duas páginas"),
    ("MÉDIO", "Mega-Sena",
     "Chamar register_conferencia_extras como as outras",
     "Paridade de conferência"),
    ("BAIXO", "Todas",
     "Unificar badge 5/6 MODELOS",
     "Cosmético"),
    ("BAIXO", "docs",
     "Atualizar verify_nav.py e o Excel antigo gerar_mapa_menus_excel.py",
     "Ferramentas defasadas"),
    ("BAIXO", "Central 8083",
     "Timeout em /configuracoes/",
     "Fora do menu das modalidades"),
]

# ---------------------------------------------------------------------------
# Estilos
# ---------------------------------------------------------------------------
thin = Side(style="thin", color="B4B4B4")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1B4F72")
title_font = Font(name="Segoe UI", bold=True, size=16, color="1B4F72")
subtitle_font = Font(name="Segoe UI", italic=True, size=10, color="666666")
body_font = Font(name="Segoe UI", size=10)
bold_font = Font(name="Segoe UI", bold=True, size=10)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

FILL = {
    "OK": PatternFill("solid", fgColor="C8E6C9"),
    "OK*": PatternFill("solid", fgColor="C8E6C9"),
    "404": PatternFill("solid", fgColor="FFCDD2"),
    "AUSENTE": PatternFill("solid", fgColor="ECEFF1"),
    "ESP": PatternFill("solid", fgColor="BBDEFB"),
    "PARCIAL": PatternFill("solid", fgColor="FFE082"),
    "CRÍTICO": PatternFill("solid", fgColor="EF9A9A"),
    "ALTO": PatternFill("solid", fgColor="FFCC80"),
    "MÉDIO": PatternFill("solid", fgColor="FFF59D"),
    "BAIXO": PatternFill("solid", fgColor="C5E1A5"),
    "SIM": PatternFill("solid", fgColor="C8E6C9"),
    "NÃO": PatternFill("solid", fgColor="FFCDD2"),
}


def style_header(ws, row, ncol, fill=None):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = center
        cell.border = border_all


def auto_width(ws, min_w=10, max_w=48):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max((len(str(c.value or "").split("\n")[0]) for c in col), default=min_w)
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def put(ws, r, c, value, *, font=None, fill=None, align=None):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = font or body_font
    cell.border = border_all
    cell.alignment = align or left_wrap
    if fill:
        cell.fill = fill
    key = str(value) if value is not None else ""
    if key in FILL and fill is None:
        cell.fill = FILL[key]
        cell.alignment = center
        cell.font = bold_font
    return cell


def http_of(key, href):
    path = (href or "/").split("#")[0] or "/"
    return HTTP_MENU.get(path, {}).get(key, 200)


# ---------------------------------------------------------------------------
# Abas
# ---------------------------------------------------------------------------
def sheet_capa(wb):
    ws = wb.active
    ws.title = "Capa"
    ws["A1"] = "Auditoria de menus e análises — modalidades 5152 a 5160"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · "
        "Fonte: nav_config.py + app.py/analise_routes.py + HTTP live · "
        "Nenhuma alteração na aplicação."
    )
    ws["A2"].font = subtitle_font
    ws.merge_cells("A2:F2")

    ws["A4"] = "Pergunta"
    ws["B4"] = "Resposta"
    style_header(ws, 4, 2)
    rows = [
        ("Qual é a modalidade mais completa?",
         "Dia de Sorte (porta 5153). O menu Análise dela é o modelo usado em nav_config.py."),
        ("Quais já estão alinhadas e funcionando?",
         "Dia de Sorte, Lotofácil, Quina, Lotomania, Mega-Sena, Timemania."),
        ("Onde o menu mente (404)?",
         "+Milionária e Dupla Sena: 3 análises. Super Sete: Escolha Visual e Análise Tubular."),
        ("O que NÃO é falta de padronização?",
         "Mês, trevos, times, 2º sorteio, colunas do Super Sete, ciclo desligado no SS, concursos especiais, Des2 da Mega."),
        ("Central",
         "Dashboard master em http://localhost:8083/ · Dia de Sorte em http://localhost:5153/"),
        ("Próxima etapa",
         "Não alterar o que já funciona. Ligar os wire_* que faltam em +M/DS; tirar do Super Sete o que não cabe em colunas; corrigir HTTP 500."),
    ]
    for i, (q, a) in enumerate(rows, 5):
        put(ws, i, 1, q, font=bold_font)
        put(ws, i, 2, a)
        ws.row_dimensions[i].height = 36
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 95

    ws["A12"] = "Abas deste arquivo"
    ws["A12"].font = Font(name="Segoe UI", bold=True, size=12, color="1B4F72")
    abas = [
        ("Capa", "Resumo executivo"),
        ("01 Modalidades", "As 9 modalidades, portas e pastas"),
        ("02 Referência DDS", "Por que Dia de Sorte é o modelo"),
        ("03 Inventário menus", "Todos os itens de menu, rota e HTTP"),
        ("04 Matriz análises", "Comparativo cruzado das análises"),
        ("05 Menu quebrado", "Item no menu sem implementação"),
        ("06 Ausentes de verdade", "Não existem no backend"),
        ("07 Nomenclatura", "Nomes atuais × nome sugerido"),
        ("08 Rotas", "Links duplicados, 404 esperados, checkers velhos"),
        ("09 Específicas", "Não padronizar (depende da regra)"),
        ("10 Prioridade", "CRÍTICO → BAIXO para a 2ª etapa"),
        ("Legenda", "Códigos da matriz"),
    ]
    ws["A13"], ws["B13"] = "Aba", "Conteúdo"
    style_header(ws, 13, 2)
    for i, (a, b) in enumerate(abas, 14):
        put(ws, i, 1, a, font=bold_font)
        put(ws, i, 2, b)


def sheet_modalidades(wb):
    ws = wb.create_sheet("01 Modalidades")
    ws["A1"] = "Modalidades efetivas no projeto"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")
    ws["A2"] = "Fonte: _shared/configuracoes/config.py · Central = 8083 · 5151 = Dia de Sorte legado (não usar)"
    ws["A2"].font = subtitle_font
    headers = ["Key", "Nome", "Porta", "Pasta", "Regra / extra", "Itens no menu Análise", "Situação das análises"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(headers))
    sit = {
        "lotofacil": "Completa (gerador Repetição → Apostas = HTTP 500)",
        "diadesorte": "REFERÊNCIA — todas as análises extras HTTP 200",
        "lotomania": "Completa e funcional",
        "quina": "Completa e funcional",
        "megasena": "Completa; referência de desdobramentos (Des1+Des2+Virada)",
        "maismilionaria": "3 análises no menu com 404; Concentração ausente; Análises Gerais sem abas",
        "duplasena": "3 análises no menu com 404; Concentração ausente; Análises Gerais sem abas",
        "timemania": "Completa e funcional",
        "supersete": "Escolha Visual e Tubular no menu com 404; ciclo desligado (específico)",
    }
    n_analise = {k: len(MENUS[k]["Análise"]) for k in KEYS}
    for i, (key, nome, porta, pasta, extra) in enumerate(MODS, 5):
        fill = PatternFill("solid", fgColor=COLORS[key])
        vals = [key, nome, porta, pasta, extra, n_analise[key], sit[key]]
        for c, v in enumerate(vals, 1):
            put(ws, i, c, v, fill=fill if c <= 5 else None, align=center if c in (3, 6) else left_wrap)
        ws.row_dimensions[i].height = 32
    auto_width(ws)
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["G"].width = 72
    ws.freeze_panes = "A5"


def sheet_referencia(wb):
    ws = wb.create_sheet("02 Referência DDS")
    ws["A1"] = "Modalidade de referência: Dia de Sorte (5153)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")
    ws["A2"] = "Usar esta estrutura na 2ª etapa. Não copiar extras matemáticos (mês, trevo, time, colunas)."
    ws["A2"].font = subtitle_font
    put(ws, 4, 1, "Por quê", font=bold_font, fill=header_fill)
    put(ws, 4, 2, "Detalhe", font=header_font, fill=header_fill)
    ws.cell(row=4, column=1).font = header_font
    motivos = [
        ("Código", "nav_config.py trata o DDS como modelo (“Paridade com Dia de Sorte”) e lista as análises extras manualmente."),
        ("HTTP", "Todas as análises extras do DDS respondem 200."),
        ("Menu Análise limpo", "Não duplica o gerador “Comportamento → Apostas” dentro de Análise (as outras volantes sim)."),
        ("Geradores", "13 itens — conjunto mais completo, origem dos geradores replicados."),
        ("Extra próprio", "Mês da Sorte em Dados — específico, não replicar."),
        ("Mega-Sena", "É a referência só de DESDOBRAMENTOS (Des1 + Des2 + Mega da Virada), não de análises."),
    ]
    for i, (a, b) in enumerate(motivos, 5):
        put(ws, i, 1, a, font=bold_font, fill=PatternFill("solid", fgColor=COLORS["diadesorte"]))
        put(ws, i, 2, b)
        ws.row_dimensions[i].height = 32

    ws["A12"] = "Estrutura-alvo do menu Análise (padrão)"
    ws["A12"].font = Font(name="Segoe UI", bold=True, size=12, color="1B4F72")
    headers = ["#", "Item padronizado", "Rota"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=13, column=i, value=h)
    style_header(ws, 13, 3)
    alvo = [
        "Sniper (nome + extra da modalidade)",
        "Resumo Geral da Modalidade",
        "Resultados & Padrões",
        "Análises Gerais",
        "Análise de Somas e Dígitos",
        "Linhas & DD × DU",
        "Análise por Posição",
        "Concentração de Acertos (se a spec existir)",
        "Análise de Ciclos das Dezenas (se motor conjunto)",
        "Análise por Gaps e Ciclo",
        "Análise Comportamental",
        "Escolha Visual (se volante de dezenas)",
        "Análise Tubular (se volante de dezenas)",
        "Central de Modelos",
        "Backtesting",
    ]
    rotas = [
        "/analise/  (LF: /analise/atrasos)",
        "/analise/resumo-geral/",
        "/analise/analises-inteligentes/",
        "/analise/analises-gerais/",
        "/analise/somas-digitos/",
        "/analise/linhas-dd-du/",
        "/analise/por-posicao/",
        "/analise/concentracao-acertos/",
        "/analise/ciclo-cobertura/",
        "/analise/gaps-ciclo/",
        "/analise/comportamento/",
        "/analise/escolha-visual/",
        "/analise/analise-tubular/",
        "/modelos/",
        "/modelos/#pane-bt",
    ]
    for i, (t, r) in enumerate(zip(alvo, rotas), 14):
        put(ws, i, 1, i - 13, align=center)
        put(ws, i, 2, t)
        put(ws, i, 3, r)
    auto_width(ws)
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 42


def sheet_inventario(wb):
    ws = wb.create_sheet("03 Inventário menus")
    ws["A1"] = "Inventário completo — barra superior de cada modalidade"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "HTTP: 200 = página responde; 404 = menu aponta para rota inexistente; "
        "500 = gerador no menu com erro. Conferência (3 âncoras) e ⚙ Configurações existem em todas."
    )
    ws["A2"].font = subtitle_font
    headers = ["Porta", "Modalidade", "Grupo", "Item do menu", "Rota", "HTTP", "Situação", "Observação"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, 8)
    r = 5
    for key, nome, porta, pasta, extra in MODS:
        fill_m = PatternFill("solid", fgColor=COLORS[key])
        for grupo, itens in MENUS[key].items():
            for titulo, href in itens:
                st = http_of(key, href)
                if st == 200:
                    sit, obs = "OK", ""
                elif st == 404:
                    sit, obs = "No menu, rota 404", "Situação B — tirar do menu ou implementar"
                elif st == 500:
                    sit, obs = "No menu, HTTP 500", "Corrigir gerador"
                else:
                    sit, obs = str(st), ""
                if titulo.startswith("Comportamento ") and href.endswith("comportamento-apostas/"):
                    obs = (obs + " · ").strip(" ·") + "Duplicata do gerador (DDS não tem isto no Análise)"
                    if sit == "OK":
                        sit = "OK (duplicata)"
                vals = [porta, nome, grupo, titulo, href, st, sit, obs]
                for c, v in enumerate(vals, 1):
                    f = fill_m if c <= 2 else None
                    put(ws, r, c, v, fill=f, align=center if c in (1, 6) else left_wrap)
                    if c == 6 and st != 200:
                        ws.cell(row=r, column=c).fill = FILL["404"] if st == 404 else FILL["PARCIAL"]
                    if c == 7 and st != 200:
                        ws.cell(row=r, column=c).fill = FILL["404"] if st == 404 else FILL["PARCIAL"]
                r += 1
        # conferência + config
        for titulo, href in [
            ("Conferência Histórica", "/central-conferencias/#historico"),
            ("Conversor de Apostas", "/central-conferencias/#conversor"),
            ("Concursos Disponíveis", "/central-conferencias/#concursos"),
            ("Configurações", "/configuracoes/"),
        ]:
            vals = [porta, nome, "Conferência / Config", titulo, href, 200, "OK", ""]
            for c, v in enumerate(vals, 1):
                put(ws, r, c, v, fill=fill_m if c <= 2 else None, align=center if c in (1, 6) else left_wrap)
            r += 1
    ws.auto_filter.ref = f"A4:H{r - 1}"
    ws.freeze_panes = "A5"
    auto_width(ws)
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 48
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 55


def sheet_matriz(wb):
    ws = wb.create_sheet("04 Matriz análises")
    ws["A1"] = "Matriz comparativa — análises × modalidades"
    ws["A1"].font = title_font
    ws.merge_cells("A1:L1")
    ws["A2"] = (
        "OK = no menu e HTTP 200 · 404 = no menu e rota quebra · AUSENTE = não está no menu nem no backend · "
        "ESP = específica da regra · PARCIAL = rota 200 sem conteúdo · OK* = equivalente com outro nome/rota"
    )
    ws["A2"].font = subtitle_font
    headers = ["Análise padronizada"] + [f"{s}\n{NOMES[k]}" for s, k in zip(SHORT, KEYS)] + ["Nota"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(headers))
    ws.row_dimensions[4].height = 36
    for i, (nome, vals, nota) in enumerate(MATRIZ, 5):
        put(ws, i, 1, nome, font=bold_font)
        for j, v in enumerate(vals, 2):
            put(ws, i, j, v, align=center)
        put(ws, i, 11, nota)
        ws.row_dimensions[i].height = 28
    ws.freeze_panes = "B5"
    auto_width(ws, min_w=12, max_w=22)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["K"].width = 62


def sheet_quebrados(wb):
    ws = wb.create_sheet("05 Menu quebrado")
    ws["A1"] = "Item aparece no menu, mas a funcionalidade não corresponde"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    headers = [
        "Modalidade", "Menu", "Análise / item", "Existe no backend?",
        "Rota existente?", "Template existente?", "Situação", "Prioridade",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 8)
    for i, row in enumerate(QUEBRADOS, 4):
        for c, v in enumerate(row, 1):
            put(ws, i, c, v)
        ws.row_dimensions[i].height = 36
    ws.freeze_panes = "A4"
    auto_width(ws)
    ws.column_dimensions["G"].width = 62


def sheet_ausentes(wb):
    ws = wb.create_sheet("06 Ausentes de verdade")
    ws["A1"] = "Não existem no backend (Situação A) — diferente de “está no menu e quebra”"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    headers = ["Análise", "Modalidades", "Evidência no código", "Ação sugerida (2ª etapa)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 4)
    for i, row in enumerate(INEXISTENTES, 4):
        for c, v in enumerate(row, 1):
            put(ws, i, c, v)
        ws.row_dimensions[i].height = 36
    auto_width(ws)
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 48


def sheet_nomes(wb):
    ws = wb.create_sheet("07 Nomenclatura")
    ws["A1"] = "Inconsistências de nome — sugestão sem alterar código nesta etapa"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")
    headers = ["Nomes atuais no projeto", "Nome padrão sugerido", "Observação"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 3)
    for i, row in enumerate(NOMENCLATURA, 4):
        for c, v in enumerate(row, 1):
            put(ws, i, c, v)
        ws.row_dimensions[i].height = 36
    auto_width(ws)
    ws.column_dimensions["A"].width = 70
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 62


def sheet_rotas(wb):
    ws = wb.create_sheet("08 Rotas")
    ws["A1"] = "Rotas, links e checkers inconsistentes"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")
    headers = ["Problema", "Onde", "Efeito"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 3)
    for i, row in enumerate(ROTAS, 4):
        for c, v in enumerate(row, 1):
            put(ws, i, c, v)
        ws.row_dimensions[i].height = 32
    auto_width(ws)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 62
    ws.column_dimensions["C"].width = 48


def sheet_especificas(wb):
    ws = wb.create_sheet("09 Específicas")
    ws["A1"] = "Não padronizar — depende da regra matemática da modalidade"
    ws["A1"].font = title_font
    ws.merge_cells("A1:C1")
    headers = ["Funcionalidade", "Modalidade", "Motivo"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 3)
    for i, row in enumerate(ESPECIFICAS, 4):
        for c, v in enumerate(row, 1):
            put(ws, i, c, v, fill=PatternFill("solid", fgColor="BBDEFB") if c == 1 else None)
        ws.row_dimensions[i].height = 32
    auto_width(ws)
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 72


def sheet_prioridade(wb):
    ws = wb.create_sheet("10 Prioridade")
    ws["A1"] = "Ordem sugerida para a 2ª etapa — preservar o que já funciona"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    headers = ["Prioridade", "Onde", "O que fazer", "Por quê (sem mexer no que está OK)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 4)
    for i, row in enumerate(PRIORIDADE, 4):
        for c, v in enumerate(row, 1):
            put(ws, i, c, v)
        ws.row_dimensions[i].height = 40
    auto_width(ws)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 78
    ws.column_dimensions["D"].width = 42


def sheet_legenda(wb):
    ws = wb.create_sheet("Legenda")
    ws["A1"] = "Legenda"
    ws["A1"].font = title_font
    linhas = [
        ("OK", "Item no menu e rota HTTP 200."),
        ("OK*", "Equivalente com outro nome ou outra rota (ex.: Lotofácil /analise/atrasos)."),
        ("404", "Item no menu; a rota não existe (wire_* ausente no analise_routes.py)."),
        ("AUSENTE", "Não está no menu e não há implementação/spec."),
        ("PARCIAL", "Rota existe (200) mas o conteúdo está vazio (0 abas em Análises Gerais)."),
        ("ESP", "Específica da modalidade — não é erro de padronização."),
        ("CRÍTICO", "Usuário clica no menu e recebe 404. Corrigir primeiro."),
        ("ALTO", "Página oca, HTTP 500 ou spec faltando com impacto real."),
        ("MÉDIO", "Nome, duplicata de menu, paridade cosmética."),
        ("BAIXO", "Badge, docs antigos, timeout da Central."),
        ("Situação A", "Não existe no projeto."),
        ("Situação B", "Existe no menu (ou deveria), implementação incompleta."),
        ("Situação C", "Mesma função com outro nome."),
        ("Situação D", "Depende da regra da modalidade."),
        ("Fonte do menu", "_shared/menu/nav_config.py → get_nav_config() + includes em templates/base.html"),
        ("Fonte das rotas", "routes/analise_routes.py (wire_*) e app.py (extend_ciclo_cobertura_app)"),
        ("HTTP", "GET em 127.0.0.1:5152–5160 em 31/08/2026. Servidores estavam no ar."),
    ]
    ws["A3"], ws["B3"] = "Código", "Significado"
    style_header(ws, 3, 2)
    for i, (a, b) in enumerate(linhas, 4):
        put(ws, i, 1, a, font=bold_font, align=center)
        put(ws, i, 2, b)
        if a in FILL:
            ws.cell(row=i, column=1).fill = FILL[a]
    auto_width(ws)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 95


def main():
    wb = Workbook()
    sheet_capa(wb)
    sheet_modalidades(wb)
    sheet_referencia(wb)
    sheet_inventario(wb)
    sheet_matriz(wb)
    sheet_quebrados(wb)
    sheet_ausentes(wb)
    sheet_nomes(wb)
    sheet_rotas(wb)
    sheet_especificas(wb)
    sheet_prioridade(wb)
    sheet_legenda(wb)
    try:
        wb.save(OUT)
        print(f"Gerado: {OUT}")
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "-novo.xlsx")
        wb.save(alt)
        print(f"Arquivo original aberto no Excel. Gerado: {alt}")


if __name__ == "__main__":
    main()
