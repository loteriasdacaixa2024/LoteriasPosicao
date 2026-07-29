# -*- coding: utf-8 -*-
"""Gera mapa-menus-modalidades-5152-5160.xlsx com formatação."""
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Instale: pip install openpyxl")

OUT = Path(__file__).parent / "mapa-menus-modalidades-5152-5160.xlsx"

MODALIDADES = [
    ("5152", "Lotofácil", "AnalisePorPosicao-Lotofacil-Only"),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only"),
    ("5154", "Lotomania", "AnalisePorPosicao-Lotomania-Only"),
    ("5155", "Quina", "AnalisePorPosicao-Quina-Only"),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only"),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only"),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only"),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only"),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only"),
]

DETALHE = [
    ("5152", "Lotofácil", "AnalisePorPosicao-Lotofacil-Only", "Dados", "Sincronização & Sorteios", "/", "Home + sync"),
    ("5152", "Lotofácil", "AnalisePorPosicao-Lotofacil-Only", "Dados", "Atrasos Posicionais", "/analise/atrasos", "Só Lotofácil"),
    ("5152", "Lotofácil", "AnalisePorPosicao-Lotofacil-Only", "Análise", "Sniper por Posição", "/analise/atrasos", "Rank vertical P1–P15"),
    ("5152", "Lotofácil", "AnalisePorPosicao-Lotofacil-Only", "Análise", "Central de Modelos (5 MODELOS)", "/modelos/", ""),
    ("5152", "Lotofácil", "AnalisePorPosicao-Lotofacil-Only", "Análise", "Backtesting Histórico", "/modelos/#pane-bt", "Âncora na página modelos"),
    ("5152", "Lotofácil", "AnalisePorPosicao-Lotofacil-Only", "—", "Conferência Histórica", "/central-conferencias/", "Link direto"),
    ("5152", "Lotofácil", "AnalisePorPosicao-Lotofacil-Only", "—", "Painel Geral", "/?painel=1", "Query painel=1"),
    ("5152", "Lotofácil", "AnalisePorPosicao-Lotofacil-Only", "Geradores de Elite", "Engine Final", "/geradores-elite/engine-final/", ""),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only", "Dados", "Sincronização & Sorteios", "/", ""),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only", "Dados", "Análise por Coluna", "/analise/", "Freq/atraso C1–C7"),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only", "Dados", "Repetição de Dígitos", "/analise/#repeticoes", "Âncora na mesma página"),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only", "Análise", "Sniper por Colunas", "/analise/", ""),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only", "Análise", "Central de Modelos (6 MODELOS)", "/modelos/", ""),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only", "Análise", "Backtesting Histórico", "/modelos/", ""),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only", "—", "Conferência Histórica", "/central-conferencias/", ""),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only", "—", "Painel Geral", "/", ""),
    ("5153", "Super Sete", "AnalisePorPosicao-SuperSete-Only", "Geradores de Elite", "Engine Final", "/geradores-elite/engine-final/", ""),
    ("5154", "Lotomania", "AnalisePorPosicao-Lotomania-Only", "Dados", "Sincronização & Sorteios", "/", ""),
    ("5154", "Lotomania", "AnalisePorPosicao-Lotomania-Only", "Dados", "Análise Estatística", "/analise/", ""),
    ("5154", "Lotomania", "AnalisePorPosicao-Lotomania-Only", "Análise", "Sniper por Dezenas", "/analise/", ""),
    ("5154", "Lotomania", "AnalisePorPosicao-Lotomania-Only", "Análise", "Central de Modelos (6 MODELOS)", "/modelos/", ""),
    ("5154", "Lotomania", "AnalisePorPosicao-Lotomania-Only", "Análise", "Backtesting Histórico", "/modelos/", ""),
    ("5154", "Lotomania", "AnalisePorPosicao-Lotomania-Only", "—", "Conferência", "/central-conferencias/", ""),
    ("5154", "Lotomania", "AnalisePorPosicao-Lotomania-Only", "—", "Painel Geral", "/", ""),
    ("5154", "Lotomania", "AnalisePorPosicao-Lotomania-Only", "Geradores de Elite", "Engine Final", "/geradores-elite/engine-final/", ""),
    ("5155", "Quina", "AnalisePorPosicao-Quina-Only", "Dados", "Sincronização & Sorteios", "/", ""),
    ("5155", "Quina", "AnalisePorPosicao-Quina-Only", "Dados", "Análise Estatística", "/analise/", ""),
    ("5155", "Quina", "AnalisePorPosicao-Quina-Only", "Análise", "Sniper por Dezenas", "/analise/", ""),
    ("5155", "Quina", "AnalisePorPosicao-Quina-Only", "Análise", "Central de Modelos (6 MODELOS)", "/modelos/", ""),
    ("5155", "Quina", "AnalisePorPosicao-Quina-Only", "Análise", "Backtesting Histórico", "/modelos/", ""),
    ("5155", "Quina", "AnalisePorPosicao-Quina-Only", "—", "Conferência", "/central-conferencias/", ""),
    ("5155", "Quina", "AnalisePorPosicao-Quina-Only", "—", "Painel Geral", "/", ""),
    ("5155", "Quina", "AnalisePorPosicao-Quina-Only", "Geradores de Elite", "Engine Final", "/geradores-elite/engine-final/", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Dados", "Sincronização & Sorteios", "/", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Dados", "Análise Estatística", "/analise/", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Análise", "Sniper por Dezenas", "/analise/", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Análise", "Central de Modelos (6 MODELOS)", "/modelos/", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Análise", "Backtesting Histórico", "/modelos/", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Desdobramentos", "Des1 — Desdobramento Inteligente", "/desdobramento/", "Só Mega"),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Desdobramentos", "Des2 — Desdobramento Estrutural", "/des2/", "Só Mega"),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Central de Conferências", "Conferência Histórica", "/central-conferencias/#historico", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Central de Conferências", "Conversor de Apostas", "/central-conferencias/#conversor", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Central de Conferências", "Concursos Disponíveis", "/central-conferencias/#concursos", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "—", "Painel Geral", "/", ""),
    ("5156", "Mega-Sena", "AnalisePorPosicao-MegaSena-Only", "Geradores de Elite", "Engine Final", "/geradores-elite/engine-final/", ""),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only", "Dados", "Sincronização & Sorteios", "/", ""),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only", "Dados", "Análise Estatística", "/analise/", "Dezenas + trevos"),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only", "Análise", "Sniper por Dezenas + Trevo", "/analise/", ""),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only", "Análise", "Central de Modelos (6 MODELOS)", "/modelos/", ""),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only", "Análise", "Backtesting Histórico", "/modelos/", ""),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only", "Desdobramento", "Desdobramento Inteligente", "/desdobramento/", "Submenu"),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only", "—", "Conferência", "/central-conferencias/", ""),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only", "—", "Painel Geral", "/", ""),
    ("5157", "+Milionária", "AnalisePorPosicao-MaisMilionaria-Only", "Geradores de Elite", "Engine Final", "/geradores-elite/engine-final/", ""),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only", "Dados", "Sincronização & Sorteios", "/", ""),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only", "Dados", "Análise Estatística", "/analise/", ""),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only", "Análise", "Sniper por Dezenas", "/analise/", ""),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only", "Análise", "Central de Modelos (6 MODELOS)", "/modelos/", ""),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only", "Análise", "Backtesting · Prêmio Duplo", "/modelos/", ""),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only", "—", "Desdobramento", "/desdobramento/", "Link direto"),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only", "—", "Conferência", "/central-conferencias/", ""),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only", "—", "Painel", "/", ""),
    ("5158", "Dupla Sena", "AnalisePorPosicao-DuplaSena-Only", "Geradores de Elite", "Engine Final", "/geradores-elite/engine-final/", ""),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only", "Dados", "Sincronização", "/", "Dezenas + time"),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only", "Dados", "Análise Estatística", "/analise/", ""),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only", "Análise", "Sniper por Dezenas + Timemania", "/analise/", ""),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only", "Análise", "Ranking de Times", "/analise/", "Mesma rota /analise/"),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only", "Análise", "Central de Modelos (6)", "/modelos/", ""),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only", "Análise", "Backtesting · 8 prêmios", "/modelos/", ""),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only", "—", "Conferência", "/central-conferencias/", ""),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only", "—", "Painel", "/", ""),
    ("5159", "Timemania", "AnalisePorPosicao-Timemania-Only", "Geradores de Elite", "Engine Final", "/geradores-elite/engine-final/", ""),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "Dados", "Sincronização", "/", ""),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "Dados", "Análise Estatística", "/analise/", ""),
    ("5153", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "Análise", "Sniper por Dezenas + Mês da Sorte", "/analise/", ""),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "Análise", "Central de Modelos (6)", "/modelos/", ""),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "Análise", "Backtesting Histórico", "/modelos/", ""),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "—", "Conferência", "/central-conferencias/", ""),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "—", "Desdobramento", "/desdobramento/", ""),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "—", "Painel", "/", ""),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "—", "Configurações", "/configuracoes/", "Só 5160"),
    ("5160", "Dia de Sorte", "AnalisePorPosicao--DiaDeSorte-Only", "Geradores de Elite", "Engine Final", "/geradores-elite/engine-final/", ""),
]

MATRIZ_ROWS = [
    ("Dados (submenu)", [1, 1, 1, 1, 1, 1, 1, 1, 1]),
    ("Atrasos posicionais", [1, 0, 0, 0, 0, 0, 0, 0, 0]),
    ("Análise por coluna / repetições", [0, 1, 0, 0, 0, 0, 0, 0, 0]),
    ("Análise (submenu: sniper + modelos)", [1, 1, 1, 1, 1, 1, 1, 1, 1]),
    ("Ranking de times", [0, 0, 0, 0, 0, 0, 0, 1, 0]),
    ("Desdobramento(s)", ["", "", "", "", "Des1+Des2", "submenu", "link", "", "link"]),
    ("Conferência (link simples)", [1, 1, 1, 1, 0, 1, 1, 1, 1]),
    ("Central de Conferências (3 itens)", [0, 0, 0, 0, 1, 0, 0, 0, 0]),
    ("Configurações", [0, 0, 0, 0, 0, 0, 0, 0, 1]),
    ("Geradores de Elite → Engine Final", [1, 1, 1, 1, 1, 1, 1, 1, 1]),
]

# Preços informados pelo usuário + concurso do site (pesquisa web em 28/05/2026).
MODALIDADES_PRECO_CONCURSO = [
    ("Mega-Sena", "R$ 6,00", 3011),
    ("Lotofácil", "R$ 3,50", 3696),
    ("Quina", "R$ 3,00", 7036),
    ("Lotomania", "R$ 3,00", 2928),
    ("Dupla Sena", "R$ 3,00", 2961),
    ("Timemania", "R$ 3,50", 2396),
    ("Dia de Sorte", "R$ 2,50", 1218),
    ("Super Sete", "R$ 3,00", 852),
    ("+Milionária", "R$ 6,00", 357),
]

# Cores por modalidade (tema suave)
COLORS = {
    "5152": "E8D7FA",
    "5153": "D4EDDA",
    "5154": "FFF3CD",
    "5155": "CCE5FF",
    "5156": "A8E6CF",
    "5157": "FFE0B2",
    "5158": "F8D7DA",
    "5159": "D1ECF1",
    "5160": "B8E0C8",
}

thin = Side(style="thin", color="B4B4B4")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2C3E50")
title_font = Font(name="Segoe UI", bold=True, size=14, color="2C3E50")
subtitle_font = Font(name="Segoe UI", italic=True, size=10, color="666666")
body_font = Font(name="Segoe UI", size=10)
check_font = Font(name="Segoe UI", bold=True, size=12, color="1E7E34")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
yes_fill = PatternFill("solid", fgColor="D4EDDA")
no_fill = PatternFill("solid", fgColor="F8F9FA")
special_fill = PatternFill("solid", fgColor="FFF3CD")


def style_header_row(ws, row, ncol, fill=None):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = center if c > 1 else left_wrap
        cell.border = border_all


def auto_width(ws, min_w=10, max_w=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max((len(str(c.value or "")) for c in col), default=min_w)
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


HEADERS = [
    "Porta",
    "Modalidade",
    "Pasta do projeto",
    "Menu principal",
    "Submenu / item",
    "Rota",
    "Observação",
]


def sheet_title(porta: str, nome: str) -> str:
    """Nome de aba Excel (máx. 31 caracteres)."""
    title = f"{porta} {nome}"
    for ch in r"[]:*?/\\":
        title = title.replace(ch, "-")
    return title[:31]


def write_tabela_modalidade(ws, porta: str, nome: str, pasta: str) -> None:
    """Mesmo layout da aba Detalhe / print Mega: 7 colunas, fundo na cor do jogo."""
    cor = COLORS.get(porta, "EEEEEE")
    fill_linha = PatternFill("solid", fgColor=cor)

    ws["A1"] = f"{porta} — {nome}"
    ws["A1"].font = Font(name="Segoe UI", bold=True, size=14, color="1A1A1A")
    ws.merge_cells("A1:G1")
    ws["A2"] = pasta
    ws["A2"].font = subtitle_font
    ws.merge_cells("A2:G2")
    ws["A3"] = "Menus (barra superior + submenus) — fonte: templates/base.html"
    ws["A3"].font = subtitle_font
    ws.merge_cells("A3:G3")

    start = 5
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=start, column=i, value=h)
    style_header_row(ws, start, len(HEADERS))

    linhas = [d for d in DETALHE if d[0] == porta]
    for idx, row in enumerate(linhas, start + 1):
        for i, val in enumerate(row, 1):
            cell = ws.cell(row=idx, column=i, value=val)
            cell.font = body_font
            cell.border = border_all
            cell.alignment = left_wrap
            cell.fill = fill_linha

    ws.freeze_panes = f"A{start + 1}"
    if linhas:
        ws.auto_filter.ref = f"A{start}:G{start + len(linhas)}"
    auto_width(ws)


def build_indice(wb):
    ws = wb.create_sheet("Índice", 0)
    ws["A1"] = "Mapa de menus — uma aba por modalidade (5152–5160)"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    start = 3
    for i, h in enumerate(["Porta", "Modalidade", "Pasta", "Aba no arquivo"], 1):
        ws.cell(row=start, column=i, value=h)
    style_header_row(ws, start, 4)
    for idx, (porta, nome, pasta) in enumerate(MODALIDADES, start + 1):
        ws.cell(row=idx, column=1, value=porta)
        ws.cell(row=idx, column=2, value=nome)
        ws.cell(row=idx, column=3, value=pasta)
        aba = sheet_title(porta, nome)
        ws.cell(row=idx, column=4, value=aba)
        for c in range(1, 5):
            ws.cell(row=idx, column=c).font = body_font
            ws.cell(row=idx, column=c).border = border_all
            if c <= 3:
                ws.cell(row=idx, column=c).fill = PatternFill(
                    "solid", fgColor=COLORS.get(porta, "EEEEEE")
                )
    auto_width(ws)


def build_abas_por_modalidade(wb):
    for porta, nome, pasta in MODALIDADES:
        ws = wb.create_sheet(sheet_title(porta, nome))
        write_tabela_modalidade(ws, porta, nome, pasta)


def build_detalhe(wb):
    ws = wb.active
    ws.title = "Detalhe menus"
    ws["A1"] = "Mapa de menus — modalidades 5152 a 5160"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")
    ws["A2"] = "Fonte: templates/base.html · Atualizado automaticamente"
    ws["A2"].font = subtitle_font
    ws.merge_cells("A2:G2")

    start = 4
    for i, h in enumerate(HEADERS, 1):
        ws.cell(row=start, column=i, value=h)
    style_header_row(ws, start, len(HEADERS))

    for idx, row in enumerate(DETALHE, start + 1):
        porta = row[0]
        port_fill = PatternFill("solid", fgColor=COLORS.get(porta, "EEEEEE"))
        alt_fill = PatternFill("solid", fgColor="F5F5F5" if idx % 2 == 0 else "FFFFFF")
        for i, val in enumerate(row, 1):
            cell = ws.cell(row=idx, column=i, value=val)
            cell.font = body_font
            cell.border = border_all
            cell.alignment = left_wrap
            cell.fill = port_fill if i <= 3 else alt_fill

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{start}:G{start + len(DETALHE)}"
    auto_width(ws)


def build_matriz(wb):
    ws = wb.create_sheet("Matriz resumida")
    ws["A1"] = "Matriz resumida — item presente no menu (✓)"
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    ws["A2"] = "Portas 5152–5160 · sem 5151"
    ws["A2"].font = subtitle_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    start = 4
    ws.cell(row=start, column=1, value="Item / rota")
    for j, (porta, nome, _) in enumerate(MODALIDADES, 2):
        cell = ws.cell(row=start, column=j, value=f"{porta}\n{nome}")
        cell.fill = PatternFill("solid", fgColor=COLORS.get(porta, "CCCCCC"))
        cell.font = Font(name="Segoe UI", bold=True, size=10, color="1A1A1A")
    style_header_row(ws, start, 1 + len(MODALIDADES), fill=header_fill)
    ws.cell(row=start, column=1).fill = header_fill
    ws.cell(row=start, column=1).font = header_font

    for r_idx, (label, vals) in enumerate(MATRIZ_ROWS, start + 1):
        ws.cell(row=r_idx, column=1, value=label).font = Font(name="Segoe UI", bold=True, size=10)
        ws.cell(row=r_idx, column=1).alignment = left_wrap
        ws.cell(row=r_idx, column=1).border = border_all
        if r_idx % 2 == 0:
            ws.cell(row=r_idx, column=1).fill = PatternFill("solid", fgColor="ECF0F1")

        for j, v in enumerate(vals, 2):
            cell = ws.cell(row=r_idx, column=j)
            cell.border = border_all
            cell.alignment = center
            if v == 1:
                cell.value = "✓"
                cell.font = check_font
                cell.fill = yes_fill
            elif v == 0:
                cell.value = ""
                cell.fill = no_fill
            else:
                cell.value = str(v)
                cell.font = Font(name="Segoe UI", size=9, color="856404")
                cell.fill = special_fill

    ws.row_dimensions[start].height = 36
    for r in range(start + 1, start + len(MATRIZ_ROWS) + 1):
        ws.row_dimensions[r].height = 28
    ws.freeze_panes = "B5"
    auto_width(ws, min_w=12, max_w=18)
    ws.column_dimensions["A"].width = 38


def build_por_modalidade(wb):
    """Todas as modalidades em uma única aba (mesmas 7 colunas das abas individuais)."""
    ws = wb.create_sheet("Todas (uma aba)")
    ws["A1"] = "Todas as modalidades — mesmo formato das abas 5152…5160"
    ws["A1"].font = title_font
    ws.merge_cells("A1:G1")

    row = 3
    for porta, nome, pasta in MODALIDADES:
        ws.cell(row=row, column=1, value=f"{porta} — {nome}")
        ws.cell(row=row, column=1).font = Font(
            name="Segoe UI", bold=True, size=12, color="1A1A1A"
        )
        fill = PatternFill("solid", fgColor=COLORS.get(porta, "CCCCCC"))
        for c in range(1, 8):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = border_all
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
        for i, h in enumerate(HEADERS, 1):
            ws.cell(row=row, column=i, value=h)
        style_header_row(ws, row, len(HEADERS))
        row += 1
        cor = COLORS.get(porta, "EEEEEE")
        fill_linha = PatternFill("solid", fgColor=cor)
        for d in DETALHE:
            if d[0] != porta:
                continue
            for i, val in enumerate(d, 1):
                cell = ws.cell(row=row, column=i, value=val)
                cell.font = body_font
                cell.border = border_all
                cell.alignment = left_wrap
                cell.fill = fill_linha
            row += 1
        row += 1
    auto_width(ws)


def build_legenda(wb):
    ws = wb.create_sheet("Legenda")
    lines = [
        ("✓", "Item presente no menu principal (barra superior ou submenu)."),
        ("Des1+Des2", "Mega-Sena: dois desdobramentos (/desdobramento/ e /des2/)."),
        ("submenu", "+Milionária: Desdobramento com dropdown."),
        ("link", "Dupla Sena e Dia de Sorte: botão direto sem dropdown."),
        ("", "Fonte: templates/base.html em cada AnalisePorPosicao-*-Only."),
        ("", "5151 não incluído — instância separada na raiz do projeto."),
    ]
    ws["A1"] = "Legenda"
    ws["A1"].font = title_font
    for i, (sym, txt) in enumerate(lines, 3):
        ws.cell(row=i, column=1, value=sym).font = check_font
        ws.cell(row=i, column=2, value=txt).font = body_font
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 70


def build_modalidades(wb):
    ws = wb.create_sheet("Modalidades")
    ws["A1"] = "Modalidades — preço, concurso do site e próximo concurso no app"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")
    ws["A2"] = "Fonte concursos: api-loterias.moleniuk.com (consulta em 28/05/2026)"
    ws["A2"].font = subtitle_font
    ws.merge_cells("A2:D2")

    headers = ["modalidade", "preco", "site", "app"]
    start = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=start, column=i, value=h)
    style_header_row(ws, start, len(headers))

    for idx, row in enumerate(MODALIDADES_PRECO_CONCURSO, start + 1):
        modalidade, preco, site = row
        app = int(site) + 1
        valores = [modalidade, preco, site, app]
        for i, val in enumerate(valores, 1):
            cell = ws.cell(row=idx, column=i, value=val)
            cell.font = body_font
            cell.border = border_all
            cell.alignment = left_wrap if i == 1 else center
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F7F9FB")

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{start}:D{start + len(MODALIDADES_PRECO_CONCURSO)}"


def export_csv_utf8_bom():
    """CSV com BOM para Excel abrir acentos corretamente (sem formatação)."""
    csv_path = OUT.with_suffix(".csv")
    lines = [
        "Porta;Modalidade;Pasta do projeto;Menu principal;Submenu / item;Rota;Observação",
    ]
    for row in DETALHE:
        lines.append(";".join(str(x).replace(";", ",") for x in row))
    csv_path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Gerado: {csv_path}")


def main():
    wb = Workbook()
    build_detalhe(wb)
    build_matriz(wb)
    build_abas_por_modalidade(wb)
    build_por_modalidade(wb)
    build_legenda(wb)
    build_modalidades(wb)
    build_indice(wb)  # primeira aba (índice de atalhos)
    try:
        wb.save(OUT)
        print(f"Gerado: {OUT}")
    except PermissionError:
        alt = OUT.with_name(OUT.stem + "-novo.xlsx")
        wb.save(alt)
        print(f"Arquivo original aberto no Excel. Gerado: {alt}")
    export_csv_utf8_bom()


if __name__ == "__main__":
    main()
